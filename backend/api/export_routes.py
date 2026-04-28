"""
Export Routes - Export attendance to Excel or CSV
"""
import io
import csv
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse

from database.db import get_connection
from api.auth_routes import get_current_user

router = APIRouter()


@router.get("/attendance/excel")
def export_excel(subject_id: int = None, date: str = None, user=Depends(get_current_user)):
    """Export attendance as Excel (.xlsx)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    conn = get_connection()
    query = """
        SELECT st.enrollment_no, st.name, st.department, sub.name as subject,
               s.date, ar.status, ar.confidence, ar.method, ar.marked_at
        FROM attendance_records ar
        JOIN students st ON st.id = ar.student_id
        JOIN attendance_sessions s ON s.id = ar.session_id
        JOIN subjects sub ON sub.id = s.subject_id
        WHERE 1=1
    """
    params = []
    if subject_id:
        query += " AND s.subject_id = ?"
        params.append(subject_id)
    if date:
        query += " AND s.date = ?"
        params.append(date)
    query += " ORDER BY s.date DESC, st.name"

    rows = conn.execute(query, params).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Header styling
    header_fill = PatternFill("solid", fgColor="1E40AF")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    headers = ["Enrollment No", "Name", "Department", "Subject", "Date", "Status", "Confidence", "Method", "Marked At"]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 18

    # Data rows
    status_colors = {"present": "D1FAE5", "absent": "FEE2E2", "late": "FEF3C7"}
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 6:  # Status column
                color = status_colors.get(str(value), "FFFFFF")
                cell.fill = PatternFill("solid", fgColor=color)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=attendance_report.xlsx"}
    )


@router.get("/attendance/csv")
def export_csv(subject_id: int = None, date: str = None, user=Depends(get_current_user)):
    """Export attendance as CSV"""
    conn = get_connection()
    query = """
        SELECT st.enrollment_no, st.name, sub.name as subject,
               s.date, ar.status, ar.marked_at
        FROM attendance_records ar
        JOIN students st ON st.id = ar.student_id
        JOIN attendance_sessions s ON s.id = ar.session_id
        JOIN subjects sub ON sub.id = s.subject_id
        WHERE 1=1
    """
    params = []
    if subject_id:
        query += " AND s.subject_id = ?"
        params.append(subject_id)
    if date:
        query += " AND s.date = ?"
        params.append(date)

    rows = conn.execute(query, params).fetchall()
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Enrollment No", "Name", "Subject", "Date", "Status", "Marked At"])
    for row in rows:
        writer.writerow(list(row))

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=attendance.csv"}
    )
